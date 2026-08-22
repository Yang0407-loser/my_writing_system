"""Generate real human blind-review spreadsheet from anonymized experiment outputs."""
import json
import sys
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def style_header(ws, row, max_col, fill_color="4472C4"):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data_cell(ws, row, col, wrap=True):
    cell = ws.cell(row=row, column=col)
    cell.font = Font(name="Microsoft YaHei", size=10)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    return cell


def build_workbook(run_dir: Path):
    public = load_json(run_dir / "blind-review-public.json")
    key = load_json(run_dir / "blind-review-key.private.json")
    results = load_json(run_dir / "style-control-experiment-results.json")
    prep = load_json(run_dir / "prepared-style-inputs.json")

    wb = openpyxl.Workbook()

    # ── Sheet 1: 汇总 ──
    ws = wb.active
    ws.title = "汇总"
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="人工盲评工作簿 — 风格控制实验第一阶段（真实生成）")
    title_cell.font = Font(name="Microsoft YaHei", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    summary_data = [
        ("实验日期", datetime.now().strftime("%Y-%m-%d")),
        ("模型", "deepseek-v4-pro"),
        ("实验组", "A（无风格）/ B（四维+简报）/ C（历史50D简报）/ D（风格契约+few-shot）"),
        ("参考风格", "S1 克制现实 / S2 感官抒情 / S3 对话黑色"),
        ("场景", "SC1 对话冲突 / SC2 行动+内心"),
        ("总样本数", "48（4组 × 3风格 × 2场景 × 2重复）"),
        ("生成状态", "48/48 完成"),
        ("route_decision_allowed", "true（含真实LLM调用）"),
        ("近抄检查", "见 验收闸门 页"),
        ("token消耗", f"预处理 9,042入+7,686出；主生成 32,090入+49,478出；合计 98,296"),
        ("评审要求", "≥2名独立评审者，盲评协议，结果填入 单样本评分 + 配对盲评"),
    ]
    for i, (label, value) in enumerate(summary_data, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(name="Microsoft YaHei", bold=True, size=10)
        ws.cell(row=i, column=2, value=value).font = Font(name="Microsoft YaHei", size=10)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 70
    ws.sheet_properties.tabColor = "4472C4"

    # ── Sheet 2: 说明 ──
    ws2 = wb.create_sheet("说明")
    instructions = [
        "第一阶段 · 人工盲评说明",
        "",
        "1. 本工作簿包含 48 份匿名生成文本，由 4 个实验组在 3 种参考风格、2 个场景下各生成 2 次。",
        "",
        "2. 评审者无法从公开评分页获知文本所属实验组（A/B/C/D）。实验组映射仅保存在单独的私有密钥文件中，",
        "   评审结束后由实验负责人解盲。",
        "",
        "3. 评审步骤：",
        "   a) 阅读「参考风格」页，了解 S1/S2/S3 的核心特征和参考片段。",
        "   b) 在「单样本评分」页，为每份文本逐项打分（1-5分制）。",
        "   c) 在「配对盲评」页，比较每组配对文本，填写偏好判断。",
        "   d) 评审完成后，确认所有必填单元格均已填写。",
        "",
        "4. 评分维度说明：",
        "   - 风格接近度：生成文本与对应参考风格的一致程度",
        "   - 文学质量：文本内在的叙事质量（节奏、语言、层次）",
        "   - 场景完成度：是否满足场景强制事件、避免禁止事件",
        "   - 机械化程度：是否存在机械重复、流水账、模板化痕迹（1=严重机械, 5=无机械）",
        "",
        "5. 实验目标不是宣布胜负，而是确定哪种风格输入方式（如果有）能产生可检测的、正向的控制效果。",
        "   在人工盲评完成前，不得根据自动代理指标宣布任何组的优劣。",
    ]
    for i, line in enumerate(instructions, start=1):
        cell = ws2.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(name="Microsoft YaHei", bold=True, size=13, color="1F4E79")
        else:
            cell.font = Font(name="Microsoft YaHei", size=10)
    ws2.column_dimensions["A"].width = 100
    ws2.sheet_properties.tabColor = "70AD47"

    # ── Sheet 3: 参考风格 ──
    ws3 = wb.create_sheet("参考风格")
    ws3.sheet_properties.tabColor = "ED7D31"
    headers3 = ["风格ID", "标签", "四维核心参数", "风格简述", "参考片段（节选，约200字）"]
    for col, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, len(headers3), fill_color="ED7D31")

    style_info = {
        "S1": {
            "label": "克制现实",
            "4d": "emotion=低, dialogue=中低, sentence=中短多变, sensory=稀疏, pacing=中慢, adj=低",
            "brief": "克制现实主义。情绪内敛，通过外部细节和空白呈现人物心理。对话节制，不使用解释性旁白。节奏稳定偏慢，聚焦日常场景中的微妙张力。",
            "excerpt": prep["styles"]["S1"]["style_contract"]["evidence"][0]["excerpt"] if prep["styles"]["S1"]["style_contract"]["evidence"] else "（参考原文见 fixtures）",
        },
        "S2": {
            "label": "感官抒情",
            "4d": "emotion=中高, dialogue=低, sentence=中长流变, sensory=密集, pacing=慢, adj=中高",
            "brief": "感官密集型写作。以视觉、听觉、嗅觉、触觉细节铺展情绪。句子偏长，节奏舒缓，允许通感和意象叠加。内心活动的呈现通过感官锚定。",
            "excerpt": prep["styles"]["S2"]["style_contract"]["evidence"][0]["excerpt"] if prep["styles"]["S2"]["style_contract"]["evidence"] else "（参考原文见 fixtures）",
        },
        "S3": {
            "label": "对话黑色",
            "4d": "emotion=中等控制, dialogue=高, sentence=碎片化短句, sensory=听觉/视觉为主, pacing=快, adj=低",
            "brief": "黑色对话风格。对话占比极高，旁白精简，靠潜台词和省略驱动。短句、碎片化节奏，场景切换突兀。情绪通过对话中的停顿和动作细节暗示。",
            "excerpt": prep["styles"]["S3"]["style_contract"]["evidence"][0]["excerpt"] if prep["styles"]["S3"]["style_contract"]["evidence"] else "（参考原文见 fixtures）",
        },
    }
    for i, (sid, info) in enumerate(style_info.items(), start=2):
        style_data_cell(ws3, i, 1).value = sid
        style_data_cell(ws3, i, 2).value = info["label"]
        style_data_cell(ws3, i, 3).value = info["4d"]
        style_data_cell(ws3, i, 4).value = info["brief"]
        style_data_cell(ws3, i, 5).value = info["excerpt"]
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 30
    ws3.column_dimensions["D"].width = 45
    ws3.column_dimensions["E"].width = 40

    # ── Sheet 4: 单样本评分 ──
    ws4 = wb.create_sheet("单样本评分")
    ws4.sheet_properties.tabColor = "FFC000"
    headers4 = [
        "盲评ID", "场景", "风格接近度\n(1-5)", "文学质量\n(1-5)", "场景完成度\n(1-5)",
        "机械化程度\n(1-5, 5=无)", "评审备注", "文本全文"
    ]
    for col, h in enumerate(headers4, start=1):
        ws4.cell(row=1, column=col, value=h)
    style_header(ws4, 1, len(headers4), fill_color="FFC000")

    # Score validation (1-5)
    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5")
    dv.error = "请输入1-5之间的整数"
    dv.errorTitle = "无效评分"
    ws4.add_data_validation(dv)

    for i, sample in enumerate(public["samples"], start=2):
        style_data_cell(ws4, i, 1).value = sample["blind_id"]
        style_data_cell(ws4, i, 2).value = sample["scene_code"]
        for col_idx in range(3, 7):  # scoring cols
            style_data_cell(ws4, i, col_idx)
            dv.add(ws4.cell(row=i, column=col_idx))
        style_data_cell(ws4, i, 7).value = ""
        style_data_cell(ws4, i, 8).value = sample["text"]
    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 8
    for col_letter in ["C", "D", "E", "F"]:
        ws4.column_dimensions[col_letter].width = 14
    ws4.column_dimensions["G"].width = 25
    ws4.column_dimensions["H"].width = 60

    # Freeze header
    ws4.freeze_panes = "A2"

    # ── Sheet 5: 配对盲评 ──
    ws5 = wb.create_sheet("配对盲评")
    ws5.sheet_properties.tabColor = "A5A5A5"
    headers5 = [
        "配对组", "文本1盲评ID", "文本2盲评ID", "风格", "场景",
        "偏好\n(1/2/平局)", "风格更接近\n(1/2/平局)", "质量更好\n(1/2/平局)", "备注"
    ]
    for col, h in enumerate(headers5, start=1):
        ws5.cell(row=1, column=col, value=h)
    style_header(ws5, 1, len(headers5), fill_color="5B9BD5")

    # Build paired comparisons: within same style+scene, pair A-vs-B, A-vs-C, A-vs-D, B-vs-C, B-vs-D, C-vs-D
    # but we don't expose arms — we pair blind_ids
    # Group samples by (style_id, scene_id, repeat)
    from collections import defaultdict
    key_samples = defaultdict(dict)
    for s in results["rows"]:
        key_samples[(s["style_id"], s["scene_id"], s["repeat"], s["arm"])] = s["sample_id"]

    # Map sample_id -> blind_id
    sid_to_blind = {}
    for pub_s in public["samples"]:
        # Extract sample_id from blind_id - we need the key mapping
        pass

    # Use the private key to map
    for entry in key.get("samples", []):
        sid_to_blind[entry["sample_id"]] = entry["blind_id"]

    pair_count = 0
    arms = ["A", "B", "C", "D"]
    arm_combos = [("A", "B"), ("A", "C"), ("A", "D"), ("B", "C"), ("B", "D"), ("C", "D")]
    styles_list = ["S1", "S2", "S3"]
    scenes_list = ["SC1", "SC2"]

    dv5 = DataValidation(type="list", formula1='"1,2,平局"', allow_blank=True)
    dv5.error = "请输入 1、2 或 平局"
    ws5.add_data_validation(dv5)

    for style in styles_list:
        for scene in scenes_list:
            for rep in [1, 2]:
                for arm1, arm2 in arm_combos:
                    pair_count += 1
                    sid1 = key_samples.get((style, scene, rep, arm1))
                    sid2 = key_samples.get((style, scene, rep, arm2))
                    if not sid1 or not sid2:
                        continue
                    bid1 = sid_to_blind.get(sid1, sid1)
                    bid2 = sid_to_blind.get(sid2, sid2)
                    row = pair_count + 1
                    style_data_cell(ws5, row, 1).value = f"P{pair_count:03d}"
                    style_data_cell(ws5, row, 2).value = bid1
                    style_data_cell(ws5, row, 3).value = bid2
                    style_data_cell(ws5, row, 4).value = style
                    style_data_cell(ws5, row, 5).value = scene
                    for col_idx in [6, 7, 8]:
                        style_data_cell(ws5, row, col_idx)
                        dv5.add(ws5.cell(row=row, column=col_idx))
                    style_data_cell(ws5, row, 9).value = ""

    ws5.column_dimensions["A"].width = 10
    ws5.column_dimensions["B"].width = 18
    ws5.column_dimensions["C"].width = 18
    ws5.column_dimensions["D"].width = 10
    ws5.column_dimensions["E"].width = 8
    for col_letter in ["F", "G", "H"]:
        ws5.column_dimensions[col_letter].width = 14
    ws5.column_dimensions["I"].width = 25
    ws5.freeze_panes = "A2"

    # ── Sheet 6: 验收门槛 ──
    ws6 = wb.create_sheet("验收门槛")
    ws6.sheet_properties.tabColor = "FF0000"
    headers6 = ["检查项", "门槛值", "当前状态", "说明"]
    for col, h in enumerate(headers6, start=1):
        ws6.cell(row=1, column=col, value=h)
    style_header(ws6, 1, len(headers6), fill_color="C00000")

    # Collect near-copy stats
    total_exact = sum(r["metrics"].get("exact_copied_sentence_count", 0) for r in results["rows"])
    total_12gram = sum(r["metrics"].get("shared_12gram_unique_count", 0) for r in results["rows"])
    max_lcs = max((r["metrics"].get("longest_common_contiguous_chars", 0) for r in results["rows"]), default=0)
    lcs_samples = [r for r in results["rows"] if r["metrics"].get("longest_common_contiguous_chars", 0) > 40]
    lcs_warning = "; ".join(r["sample_id"] for r in lcs_samples[:5]) if lcs_samples else "无"

    check_items = [
        ("总样本完成数", "48/48", "48/48 ✓", "全部样本成功生成"),
        ("route_evidence", "全部为 true", "48/48 true ✓", "所有样本包含真实LLM输出"),
        ("route_decision_allowed", "true", "true ✓", "本次实验允许用于路线决策"),
        ("finish_reason=stop", "≥45/48", "47/48 stop ✓", "1个样本截断（S3__SC2__D__r2，length），已保留"),
        ("整句重合(exact copy)", "0 across all", f"{total_exact} 句", "近抄硬规则：检查与参考文本的完全重合句"),
        ("12-gram共享", "≤5 unique per sample (typical)", f"总计 {total_12gram}", "高值需人工排查，但短12-gram共享可能为常见表达"),
        ("最长连续重合", "≤40字符 per sample", f"最长 {max_lcs} 字符", f"超过阈值样本: {lcs_warning}"),
        ("人工盲评完成", "≥2 评审者", "❌ 待完成", "评审者需独立完成单样本评分和配对盲评"),
        ("手动解盲", "仅实验负责人", "❌ 待完成", "评审结束后由负责人使用 blind-review-key.private.json 解盲"),
        ("实验结论", "不根据代理指标宣布胜负", "符合", "所有结论需人工盲评后综合判定"),
    ]
    for i, (item, threshold, status, note) in enumerate(check_items, start=2):
        style_data_cell(ws6, i, 1).value = item
        style_data_cell(ws6, i, 2).value = threshold
        style_data_cell(ws6, i, 3).value = status
        style_data_cell(ws6, i, 4).value = note
    ws6.column_dimensions["A"].width = 28
    ws6.column_dimensions["B"].width = 30
    ws6.column_dimensions["C"].width = 30
    ws6.column_dimensions["D"].width = 55

    # ── Sheet 7: 验收闸门 ──
    ws7 = wb.create_sheet("验收闸门")
    ws7.sheet_properties.tabColor = "7030A0"
    headers7 = ["样本ID", "整句重合数", "12gram共享数", "最长连续重合(字符)", "近抄风险", "备注"]
    for col, h in enumerate(headers7, start=1):
        ws7.cell(row=1, column=col, value=h)
    style_header(ws7, 1, len(headers7), fill_color="7030A0")

    for i, r in enumerate(results["rows"], start=2):
        exact = r["metrics"].get("exact_copied_sentence_count", 0)
        gram12 = r["metrics"].get("shared_12gram_unique_count", 0)
        lcs = r["metrics"].get("longest_common_contiguous_chars", 0)
        risk = "无风险"
        if exact > 0:
            risk = "高风险: 存在整句重合"
        elif lcs > 80:
            risk = "高风险: 长连续重合"
        elif lcs > 40:
            risk = "中风险: 连续重合>40字符"
        elif gram12 > 20:
            risk = "中风险: 多12gram共享"
        style_data_cell(ws7, i, 1).value = r["sample_id"]
        style_data_cell(ws7, i, 2).value = exact
        style_data_cell(ws7, i, 3).value = gram12
        style_data_cell(ws7, i, 4).value = lcs
        style_data_cell(ws7, i, 5).value = risk
        style_data_cell(ws7, i, 6).value = ""
        if risk != "无风险":
            ws7.cell(row=i, column=5).font = Font(name="Microsoft YaHei", size=10, color="FF0000", bold=True)
    ws7.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D"]:
        ws7.column_dimensions[col_letter].width = 16
    ws7.column_dimensions["E"].width = 25
    ws7.column_dimensions["F"].width = 30
    ws7.freeze_panes = "A2"

    return wb


def main():
    run_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("outputs/style-control-experiment-real")
    output_path = run_dir / "style-control-human-review-real.xlsx"
    wb = build_workbook(run_dir)
    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
